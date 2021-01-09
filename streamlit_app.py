import openpyxl
from openpyxl import Workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
import os.path
from openpyxl import load_workbook
import datetime as dt
import streamlit as st

st.markdown("<h1 style='text-align: center; color: #08081f;'>TrackIt!</h1>", unsafe_allow_html=True)
st.markdown("<h3 style='text-align: center; color: grey;'>Financial Report Generator</h3>", unsafe_allow_html=True)
user_input = st.selectbox('What would you like to do?', ('Create new file', 'Update existing file'))
x = dt.datetime.now()
date = x.strftime("%x")

# Defines a Boolean variable, 'run'
run = True



def collect_data(numrows):
    lst = []
    mainlst = []
    #defines a count variable for unique keys
    count = 1

    # Recursion 'rows' number of times to ask user for input
    for i in range(rows):
        lst.append(date)
        user_input_val_1 = st.number_input('Enter value 1: ', key = count)
        user_input_val_2 = st.number_input('Enter value 2: ', key = count+1)
        user_input_val_3 = st.number_input('Enter value 3: ', key = count+2)
        user_input_val_4 = st.number_input('Enter value 4: ', key = count+3)
        user_input_val_5 = st.number_input('Enter value 5: ', key = count+4)
        count = count + 5
        v1 = user_input_val_1
        lst.append(v1)
        v2 = user_input_val_2
        lst.append(v2)
        v3 = user_input_val_3
        lst.append(v3)
        v4 = user_input_val_4
        lst.append(v4)
        v5 = user_input_val_5
        lst.append(v5)
                
        mainlst.append(lst)
        lst = []

    return mainlst

# Creates a while loop as long as 'run' is true...do the following
while run:


    # If user entered 'n', it means they wish to create a new Excel File
    if user_input == 'Create new file':
        
        # Asks user for name they wish to give Excel file
        user_file_input = st.text_input('What do you want to name your file?: ', key ='file_name')
        file_input  = str(user_file_input)

        filename = ''.join((file_input, '.xlsx'))

        # Creates an Excel file of name, 'filename'
        file_path = os.path.isfile(filename)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Sheet'
        ws['A1'] = 'Date'
        ws['B1'] = 'Val 1'
        ws['C1'] = 'Val 2'
        ws['D1'] = 'Val 3'
        ws['E1'] = 'Val 4'
        ws['F1'] = 'Val 5'

        col = 65
        row = '2'

        # Asks user for number of rows they wish to create
        row_input = st.number_input('How many rows do you need?: ', key = 'row_input_1')
        rows = int(row_input)

        information = collect_data(rows)

        for row in information:
            ws.append(row)

        if st.button("Save"):
            wb.save(filename = filename)
            run = False
            st.write('Successfully saved! you may exit the program.')

    # If user input is 'u', it means they wish to update an Excel File
    elif user_input == 'Update existing file':

        update_target_name = st.text_input('What is the name of the file you wish to update?: ', key = 'existing_name')
        str_convert = str(update_target_name)
        filetoupdate = ''.join((str_convert, '.xlsx'))

        wb = load_workbook(filetoupdate)

        if wb:
            page = wb.active

            row_input = st.number_input('How many rows do you need?: ', key = 'row-input_2')
            rows = int(row_input)

            update_file = collect_data(rows)

            for info in update_file:
                page.append(info)

            if st.button("Save"):
                wb.save(filename = filetoupdate)
                run = False
        
        else: 
            print('File Not Found')
