import openpyxl
from openpyxl import Workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
import os.path
from openpyxl import load_workbook
import datetime as dt

# **************************************************************************************************

# In the elif command (when the user enters 'u'), we need to return error message if 
# file does not exist with that name. The way I have done it rn doesn't work 100% because Python
# returns the error but we dont to just end the program...cause its possible user made a typo.
# **************************************************************************************************

x = dt.datetime.now()
date = x.strftime("%x")

# Defines a Boolean variable, 'run'
run = True

def collect_data(numrows):
    lst = []
    mainlst = []

    # Recursion 'rows' number of times to ask user for input
    for i in range(rows):
        lst.append(date)
        v1 = int(input('\nEnter value 1: '))
        lst.append(v1)
        v2 = int(input('Enter value 2: '))
        lst.append(v2)
        v3 = int(input('Enter value 3: '))
        lst.append(v3)
        v4 = int(input('Enter value 4: '))
        lst.append(v4)
        v5 = int(input('Enter value 5: '))
        lst.append(v5)
                
        mainlst.append(lst)
        lst = []

    return mainlst

# Creates a while loop as long as 'run' is true...do the following
while run:

    user_input = str(input('Do you want to create a new file, update an existing file or exit? (n/u/e): '))

    # If user entered 'n', it means they wish to create a new Excel File
    if user_input == 'n':
        
        # Asks user for name they wish to give Excel file
        file_input = str(input('What do you want to name your file?: '))

        filename = ''.join((file_input, '.xlsx'))

        # Creates an Excel file of name, 'filename'
        file_path = os.path.isfile(filename)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Sheet'
        ws['A1'] = 'Date'
        ws['B1'] = 'Val 1'
        ws['C1'] = 'Val 2'
        ws['D1'] = 'Val 3'
        ws['E1'] = 'Val 4'
        ws['F1'] = 'Val 5'

        col = 65
        row = '2'

        # Asks user for number of rows they wish to create
        rows = int(input('How many rows do you need?: '))

        information = collect_data(rows)

        for row in information:
            ws.append(row)

        wb.save(filename = filename)
        run = False

    # If user input is 'u', it means they wish to update an Excel File
    elif user_input == 'u':

        filetoupdate = ''.join((str(input('What is the name of the file you wish to update?: ')), '.xlsx'))

        wb = load_workbook(filetoupdate)

        if wb:
            page = wb.active

            rows = int(input('How many rows do you need?: '))

            update_file = collect_data(rows)

            for info in update_file:
                page.append(info)

            wb.save(filename = filetoupdate)
            run = False
        
        else: 
            print('File Not Found')

    # If user input is 'e', it means they wish to exit
    elif user_input == 'e':
        print('\n')
        break
        
    # If user input is 'e', it means they wish to exit
    else: 
        print('Please enter either n, u or e \n')